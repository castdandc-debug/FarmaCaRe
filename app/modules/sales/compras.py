# -*- coding: utf-8 -*-
from flask import Blueprint, render_template, request, flash, redirect, url_for, make_response
from app.models import db, Compra, Proveedor
from app.models import Medicamento, Inventario
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime

compras_bp = Blueprint('compras', __name__)

@compras_bp.route('/generar_pedido_excel')
def generar_pedido_excel():
    # Obtener productos con stock menor al punto de reorden
    # Usando un JOIN entre Medicamento e Inventario
    productos_faltantes = db.session.query(
        Medicamento, Inventario
    ).join(
        Inventario, 
        (Inventario.producto_id == Medicamento.id) & 
        (Inventario.tipo_producto == 'medicamento')
    ).filter(
        Medicamento.stock < Inventario.punto_reorden
    ).order_by(Medicamento.nombre_comercial).all()
    
    # Si no hay datos en Inventario, usar stock mínimo por defecto
    if not productos_faltantes:
        # Usar un punto de reorden por defecto de 10 para productos con stock bajo
        productos_con_stock_bajo = Medicamento.query.filter(
            Medicamento.stock < 10
        ).order_by(Medicamento.nombre_comercial).all()
        
        productos_faltantes = [(med, None) for med in productos_con_stock_bajo]
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido de Compra"
    
    # Configurar encabezados
    headers = [
        'Código de Barras',
        'Nombre Comercial', 
        'Nombre Genérico',
        'Presentación',
        'Cantidad en Inventario',
        'Punto de Reorden',
        'Cantidad Faltante'
    ]
    
    # Estilo para encabezados
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos
    for row, (medicamento, inventario) in enumerate(productos_faltantes, 2):
        # Usar datos del inventario si existe, sino usar valores por defecto
        punto_reorden = inventario.punto_reorden if inventario else 10
        cantidad_faltante = punto_reorden - medicamento.stock
        
        ws.cell(row=row, column=1, value=medicamento.codigo_barras)
        ws.cell(row=row, column=2, value=medicamento.nombre_comercial)
        ws.cell(row=row, column=3, value=medicamento.nombre_generico or '')
        ws.cell(row=row, column=4, value=medicamento.presentacion or '')
        ws.cell(row=row, column=5, value=medicamento.stock)
        ws.cell(row=row, column=6, value=punto_reorden)
        ws.cell(row=row, column=7, value=cantidad_faltante)
    
    # Ajustar ancho de columnas
    column_widths = [20, 30, 30, 25, 18, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=pedido_compra_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Mensaje flash para informar al usuario
    if not productos_faltantes:
        flash('No hay productos con faltante de inventario.', 'info')
        return redirect(url_for('compras.lista'))
    else:
        flash(f'Pedido generado con {len(productos_faltantes)} productos.', 'success')
    
    return response

@compras_bp.route('/compras')
def lista():
    compras = Compra.query.all()
    return render_template('compras.html', compras=compras)

@compras_bp.route('/compras/nueva', methods=['GET', 'POST'])
def nueva():
    proveedores = Proveedor.query.all()
    if request.method == 'POST':
        proveedor_id = int(request.form['proveedor_id'])
        items = []
        total = 0
        for key, value in request.form.items():
            if key.startswith('med_') and value:
                med_id = int(key.replace('med_', ''))
                cantidad = int(value)
                med = Medicamento.query.get(med_id)
                if med:
                    items.append((med, cantidad))
                    total += med.precio_venta * cantidad
                else:
                    flash(f'Medicamento no encontrado.')
                    return redirect(url_for('compras.nueva'))

        # Crear compra
        compra = Compra(proveedor_id=proveedor_id, total=total, usuario_id=1)
        db.session.add(compra)
        db.session.commit()

        # Actualizar stock
        for med, cantidad in items:
            med.stock += cantidad
        db.session.commit()

        flash('Compra registrada con éxito.')
        return redirect(url_for('compras.lista'))

    return render_template('nueva_compra.html', proveedores=proveedores)
